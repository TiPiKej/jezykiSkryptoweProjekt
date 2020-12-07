from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import re


def search():
    print("SEARCHING")
    file_name = input("Wpisz nazwe pliku (lokalizacje) do przeszukania: ")
    search_query = input("Podaj przeszukiwane slowo: ")

    result_arr = []
    wb = Workbook()
    # cells = {}
    try:
        wb = load_workbook(file_name, read_only=True)
        sheets = [wb.worksheets[i].title for i in range(len(wb.worksheets))]
        ws = 0

        # pobieranie nazwy arkusza do przeszukiwania
        while True:
            print("Podaj arkusz (dostepne: {}) w ktorym ma byc przeszukiwane slowo: {}"
                  .format(sheets, search_query))
            temp_sheet = input()
            if temp_sheet in sheets:
                ws = wb.worksheets[sheets.index(temp_sheet)]
                break
            print("Niepoprawna nazwa arkusza!")

        for i in range(1, ws.max_row + 1):
            for j in range(1, ws.max_column + 1):
                temp_value = ws.cell(row=i, column=j).value
                print(type(temp_value))
                if re.search(search_query, temp_value):  # wyrazenie regularne (wyrazenie, sprawdzana wartosc)
                    result_arr.append({
                        "row": i,
                        "column": j,
                        "value": temp_value
                    })
    except FileNotFoundError:
        print("File not found!!")
    except InvalidFileException:
        print("File format not supported!!")
    finally:
        wb.close()
    print(result_arr)
