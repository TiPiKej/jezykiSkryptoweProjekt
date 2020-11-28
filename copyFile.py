from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def copy_file():
    print("KOPIOWANIE PLIKU")
    file_name_src = input("Wpisz nazwe pliku (lokalizacje) do skopiowania: ")
    file_name_dst = input("Wpisz nazwe pliku (lokalizacje) gdzie skopiowac: ")
    wb1 = Workbook()
    wb2 = Workbook()

    try:
        wb1 = load_workbook(file_name_src)  # zaladowanie arkusza
        ws2 = wb2.active  # ustawienie aktywnego arkusza
        ws2.title = wb1.worksheets[0].title  # przepisanie tytulu pierwszego arkusza

        for n in range(len(wb1.worksheets)):
            ws1 = wb1.worksheets[n]  # ustawienie aktywnego arkusza (do skopiowania) na kolejny
            # kopiowanie wartosci komorek
            for i in range(1, ws1.max_row + 1):
                for j in range(1, ws1.max_column + 1):
                    cell = ws1.cell(row=i, column=j)

                    ws2.cell(row=i, column=j).value = cell.value
            # ustawianie tytulow pozostalych arkuszy
            if n + 1 < len(wb1.worksheets):
                ws2 = wb2.create_sheet(wb1.worksheets[n + 1].title)

        wb2.save(str(file_name_dst))
        print("PLIK SKOPIOWANO")
    except (InvalidFileException, PermissionError, FileNotFoundError) as e:
        print(e)
    finally:
        wb1.close()
        wb2.close()

    return 0
