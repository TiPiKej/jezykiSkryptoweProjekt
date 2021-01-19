from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from components.basicFunctions import create_reg_exp_query, print_table, get_data, get_sheet
from components.jsonInputOutput import save_to_json
from components.basicConstants import confirm_query, reg_file_name_xlsx
import re


def search():
    file_name = input("Wpisz nazwe pliku (lokalizacje) do przeszukania: ")
    if not re.search(reg_file_name_xlsx, file_name):
        file_name = "{}.xlsx".format(file_name)

    wb = Workbook()
    try:
        wb = load_workbook(file_name, read_only=True)
        search_query = create_reg_exp_query()

        # pobieranie nazwy arkusza do przeszukiwania
        ws = get_sheet(wb, search_query)

        # pobranie danych z arkusza pasujace do query
        result_arr = get_data(ws, search_query)

        # wyswietlenie tabeli w konsoli
        print_table(result_arr)

        # zapisywanie wynikow wyszukiwania do pliku
        if len(result_arr) > 0:
            choice = input("Czy zapisac wynik wyszukiwania? (t - tak): ")
            if re.search(confirm_query, choice):
                save_to_json({
                    ws.title: result_arr
                })
    except InvalidFileException:
        print("Nieobslugiwany format pliku!")
    except PermissionError as e:
        print("Brak uprawnien do pliku!", e.filename)
    except FileNotFoundError as e:
        print("Nie znaleziono lokalizacji!", e.filename)
    finally:
        wb.close()
