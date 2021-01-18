from components.basicConstants import confirm_query, reg_file_name_json, reg_file_name_xlsx
from components.basicFunctions import add_sheet_to_workbook, check_if_file_exist, get_data
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import json
import re


def save_to_json(table):
    file_name = input("Podaj nazwe pliku (lokalizacje) gdzie zapisac: ")

    if not re.search(reg_file_name_json, file_name):
        file_name = "{}.json".format(file_name)

    if check_if_file_exist(file_name):
        return "Nie zapisano pliku"

    # saving file
    try:
        with open("{}".format(file_name), "w") as json_file:
            json.dump(table, json_file)
            print("Liste zapisano w pliku {}".format(file_name))
            return "Plik zapisano"
    except IOError as e:
        print("Nie mozna zapisac pliku: ", e)
        if re.search(confirm_query, input("Czy chcesz sprobowac zapisac raz jeszcze? (t - tak) ")):
            return save_to_json(table)


def save_xlsx_to_json():
    file_name = input("Wpisz nazwe pliku (lokalizacje pliku z rozszerzeniem xlsx) do eksportu (do pliku json): ")
    result_dic = {}
    wb = Workbook()
    try:
        if not re.search(reg_file_name_xlsx, file_name):
            file_name = "{}.xlsx".format(file_name)
        wb = load_workbook(file_name, read_only=True)

        for ws in wb.worksheets:
            result_dic[ws.title] = get_data(ws)

        save_to_json(result_dic)
    except FileNotFoundError:
        print("File not found!!")
    except InvalidFileException:
        print("File format not supported!!")
    finally:
        wb.close()


def save_json_to_xlsx():
    file_name = input("Wpisz nazwe pliku (lokalizacje pliku z rozszerzeniem json) do eksportu (do pliku xlsx): ")
    if not re.search(reg_file_name_json, file_name):
        file_name = "{}.json".format(file_name)

    try:
        with open("{}".format(file_name), "r") as json_file:
            tables = json.load(json_file)
            wb = Workbook()

            file_name_out = input("Wpisz nazwe pliku gdzie zapisac plik: ")
            if not re.search(reg_file_name_xlsx, file_name_out):
                file_name_out = "{}.xlsx".format(file_name_out)

            overwrite_first_sheet = False
            if check_if_file_exist(file_name_out, True):
                wb = load_workbook(file_name_out)
            else:
                overwrite_first_sheet = True

            if type(tables) == list:
                add_sheet_to_workbook(tables, wb, overwrite=overwrite_first_sheet)
            elif type(tables) == dict:
                for table_keys in tables.keys():
                    add_sheet_to_workbook(tables[table_keys], wb, table_keys, overwrite_first_sheet)
                    overwrite_first_sheet = False
            del overwrite_first_sheet

            wb.save(file_name_out)
    except IOError as e:
        print("Nie mozna wczytac pliku!")
        print(e)
    except UnicodeDecodeError:
        print("Cos poszlo nie tak! :(")
