import random
import re
import pathlib

from typing import List
from openpyxl import Workbook
from prettytable import PrettyTable

from components.basicConstants import confirm_query, reg_file_name_xlsx


def check_if_file_exist(file_name, quiet=False):
    """

    :param file_name: string
    :param quiet: boolean; if is True -> nothing has prompt
    :return: if exist -> True, if not exist -> False
    """
    file = pathlib.Path(file_name)
    if file.exists():
        if quiet:
            return True

        confirm = input("Plik istnieje, czy chcesz go nadpisac? (t - tak): ")
        if not re.search(confirm_query, confirm):
            return True
    return False


def create_reg_exp_query():
    """
    funkcja pobierajaca wyrazenie regularne od uzytkownika

    :return: reg exp
    """

    print("""Opcje:
        1 - podstawowe wyszukiwanie (mozna uzyc wyrazenia regularnego)
        2 - wyszukiwanie zaawansowane z pomoca""")

    while True:
        choice = input("Wybor: ")
        if re.search("^[12]$", choice):
            break
        print("Podales bledna wartosc!")

    choice = int(choice)

    if choice == 1:
        return input("Podaj szukane haslo: ")

    # zaawansowane wyrazenie query
    query = ""

    if re.search(confirm_query, input("Czy mam szukac slow zaczynajacych sie od wyrazenia? (t - tak) ")):
        query = "^" + query

    query += input("Podaj zestaw znakow (np. [a-e]), lub szukane slowo (np. abc): ")

    while True:
        n = input("Ile mam szukac wystapien (np. 4), \
jezeli ma zawierac conajmniej x znakow dodaj po tej liczbie przecinek (np. 4,): ")
        if re.search("^[0-9]+[,]*$", n):
            break
        print("Podales nieprawidlowe dane!")
    query += "{" + n + "}"

    if re.search(confirm_query, input("Czy mam szukac slow konczoncych sie wyrazeniem? (t - tak) ")):
        query += "$"

    return query


def print_table(table: List[dict], *titles: str):
    """
    funkcja ktora wyswietla w konsoli dictionary

    :param table: lista dictonary
    :param titles: opcjonalny parmetr, parametry, z dictionary, do wyswietlenia
    """
    # if titles are empty -> titles are filles by first row of table
    if len(table) == 0:
        print("Brak wynikow!")
        return None

    if len(titles) == 0:
        titles = table[0].keys()

    pretty_table = PrettyTable()

    pretty_table.field_names = [title for title in titles]  # add table's header

    for cell in table:
        pretty_table.add_row([cell[title] for title in titles])

    print(pretty_table)
    print("Liczba znalezionych elementow: {}".format(len(table)))


def add_sheet_to_workbook(table: List[dict], wb=Workbook(), title="Untitled", overwrite=False):
    """
    dodaj arkusz na podstawie listy z dictionary, parametry: (row, column, value)

    :param table: list[dict] - dane do zapisania do arkusza
    :param wb: workbook - plik gdzie ma byc zapisany wynik
    :param title: string - nazwa arkusza do zapisania wyniku
    :param overwrite: boolean - jesli prawda -> to nadpisze arkusz, jesli falsz -> doda nowy arkusz
    (parametr pomocny gdy jest utworzony nowy plik i chcemy nadpisac domyslny arkusz)
    :return: nowy worksheet
    """
    if overwrite:
        ws = wb.worksheets[-1]
        ws.title = title
    else:
        ws = wb.create_sheet(title)
    # kopiowanie wartosci komorek
    for cell in table:
        ws.cell(row=cell['row'], column=cell['column']).value = cell['value']
    return ws


def new_static_file():
    wb = Workbook()
    file_name = input("Podaj nazwe pliku: ")
    if not re.search(reg_file_name_xlsx, file_name):
        file_name = "{}.xlsx".format(file_name)

    if check_if_file_exist(file_name):
        return "Nie utworzono pliku"

    ws = wb.active
    ws.title = "Tabliczka mnozenia"

    for row in range(1, 11):
        for column in range(1, 11):
            ws.cell(row=row, column=column).value = row * column

    ws = wb.create_sheet("pi")
    ws["A1"] = 3.14

    ws = wb.create_sheet("e")
    ws["A1"] = 2.718

    ws = wb.create_sheet("losowe liczby")

    for i in range(1, 100):
        ws["A{}".format(i)] = i
        ws["B{}".format(i)] = round(random.random() * 100000) / 1000

    wb.save(file_name)


def create_empty_file():
    file_name = input("Podaj nazwe pliku: ")
    if not re.search(reg_file_name_xlsx, file_name):
        file_name = "{}.xlsx".format(file_name)

    if check_if_file_exist(file_name):
        return "Nie utworzono pliku"

    wb = Workbook()
    wb.save(file_name)


def get_data(ws, search_query="^.*$"):
    """
    zwraca dane z arkusza
    :param ws: arkusz
    :param search_query: regexp
    :return: arr:
            jesli search_query ustawione -> zwraca pasujace wyniki,
            w przeciwnym wypadku -> zwraca wszystko
    """
    result_arr = []
    for i in range(ws.min_row, ws.max_row + 1):
        for j in range(ws.min_column, ws.max_column + 1):
            temp_value = ws.cell(row=i, column=j).value
            # sprawdzam czy komorka jest pusta
            if temp_value is None:
                continue
            # wyrazenie regularne (wyrazenie, sprawdzana wartosc)
            # jesli sie zgadza -> dodaje do zwracanej petli
            if re.search(search_query, str(temp_value)):
                result_arr.append({
                    "row": i,
                    "column": j,
                    "value": temp_value
                })
    return result_arr


def get_sheet(wb: Workbook, search_query=""):
    """
    wybiera arkusz
    :param wb: workbook - plik
    :param search_query: string - jesli cos jest przeszukiwane to pokaze query
    :return: arkusz wybrany przez uzytkownika
    """
    sheets = [wb.worksheets[i].title for i in range(len(wb.worksheets))]
    while True:
        if search_query == "":
            temp_sheet = input("Podaj arkusz (dostepne: {}): ".format(sheets))
        else:
            temp_sheet = input(
                "Podaj arkusz (dostepne: {}) w ktorym ma byc przeszukiwane wyrazenie ({}): ".format(sheets,
                                                                                                    search_query))
        if temp_sheet in sheets:
            ws = wb.worksheets[sheets.index(temp_sheet)]
            return ws
        print("Niepoprawna nazwa arkusza!")
