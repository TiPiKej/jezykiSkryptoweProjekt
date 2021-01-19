from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from components.basicConstants import reg_file_name_xlsx
import re


def copy_file():
    file_name_src = input("Wpisz nazwe pliku (lokalizacje) do skopiowania: ")
    file_name_dst = input("Wpisz nazwe pliku (lokalizacje) gdzie skopiowac: ")
    wb1 = Workbook()
    wb2 = Workbook()

    if not re.search(reg_file_name_xlsx, file_name_src):
        file_name_src = "{}.xlsx".format(file_name_src)

    if not re.search(reg_file_name_xlsx, file_name_dst):
        file_name_dst = "{}.xlsx".format(file_name_dst)

    print("KOPIOWANIE PLIKU")

    try:
        wb1 = load_workbook(file_name_src)  # zaladowanie arkusza
        ws2 = wb2.active  # ustawienie aktywnego arkusza
        ws2.title = wb1.worksheets[0].title  # przepisanie tytulu pierwszego arkusza

        for n in range(len(wb1.worksheets)):
            ws1 = wb1.worksheets[n]  # ustawienie aktywnego arkusza (do skopiowania) na kolejny
            # kopiowanie wartosci komorek
            for i in range(ws1.min_row, ws1.max_row + 1):
                for j in range(ws1.min_column, ws1.max_column + 1):
                    cell = ws1.cell(row=i, column=j)

                    ws2.cell(row=i, column=j).value = cell.value
            # ustawianie tytulow pozostalych arkuszy
            if n + 1 < len(wb1.worksheets):
                ws2 = wb2.create_sheet(wb1.worksheets[n + 1].title)

        wb2.save(str(file_name_dst))
        print("PLIK SKOPIOWANO")
    except InvalidFileException:
        print("Nieobslugiwany format pliku!")
    except PermissionError as e:
        print("Brak uprawnien do pliku!", e.filename)
    except FileNotFoundError as e:
        print("Nie znaleziono lokalizacji!", e.filename)
    finally:
        wb1.close()
        wb2.close()

    return 0
